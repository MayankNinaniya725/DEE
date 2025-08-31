# extractor/views.py
import os
import json
import logging
from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
from django.db.models import Count
from django.core.exceptions import ValidationError
from django.conf import settings
from django.http import JsonResponse, HttpResponse, FileResponse
from django.utils import timezone
from celery.result import AsyncResult
import pandas as pd
from openpyxl import load_workbook

from .models import ExtractedData, Vendor, UploadedPDF
from .utils.extractor import extract_pdf_fields
from .utils.config_loader import load_vendor_config
from .tasks import process_pdf_file

# Setup logging
import logging.handlers

# Create logs directory if it doesn't exist
log_dir = os.path.join(settings.BASE_DIR, 'logs')
os.makedirs(log_dir, exist_ok=True)

# Configure logging
logger = logging.getLogger('extractor')
logger.setLevel(logging.DEBUG)

# File handler for all logs
file_handler = logging.handlers.RotatingFileHandler(
    os.path.join(log_dir, 'extractor.log'),
    maxBytes=1024*1024,  # 1MB
    backupCount=5
)
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
))
if not any(isinstance(h, logging.handlers.RotatingFileHandler) and h.baseFilename == file_handler.baseFilename for h in logger.handlers):
    logger.addHandler(file_handler)

# Error file handler
error_handler = logging.handlers.RotatingFileHandler(
    os.path.join(log_dir, 'errors.log'),
    maxBytes=1024*1024,  # 1MB
    backupCount=5
)
error_handler.setLevel(logging.ERROR)
error_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
))
if not any(isinstance(h, logging.handlers.RotatingFileHandler) and h.baseFilename == error_handler.baseFilename for h in logger.handlers):
    logger.addHandler(error_handler)

print("extractor.views module loaded - Logging configured")

def dashboard(request):
    # Clean up any stale task IDs in the session
    if 'last_task_id' in request.session:
        try:
            # Check if task is still valid
            task_id = request.session['last_task_id']
            res = AsyncResult(task_id)
            # If task doesn't exist or is in a final state, remove it
            if res.state in ['SUCCESS', 'FAILURE'] or res.state is None:
                del request.session['last_task_id']
                request.session.modified = True
        except Exception:
            # Task ID is invalid, remove it
            del request.session['last_task_id']
            request.session.modified = True
    
    backups_dir = os.path.join(settings.MEDIA_ROOT, "backups")
    master_path = os.path.join(backups_dir, "master.xlsx")

    data = []
    if os.path.exists(master_path):
        # Read all sheets from Excel file
        df = pd.read_excel(master_path, sheet_name=None)
        
        # Process each sheet and convert to records
        for sheet_name, sheet_df in df.items():
            # Ensure column names match exactly with Excel file
            if 'PLATE_NO' not in sheet_df.columns and 'Plate No' in sheet_df.columns:
                sheet_df = sheet_df.rename(columns={'Plate No': 'PLATE_NO'})
            if 'HEAT_NO' not in sheet_df.columns and 'Heat No' in sheet_df.columns:
                sheet_df = sheet_df.rename(columns={'Heat No': 'HEAT_NO'})
            if 'TEST_CERT_NO' not in sheet_df.columns and 'Test Cert No' in sheet_df.columns:
                sheet_df = sheet_df.rename(columns={'Test Cert No': 'TEST_CERT_NO'})
                
            # Handle NaN values to prevent template errors
            sheet_df = sheet_df.fillna("")
            
            # Convert to records with all fields processed
            records = sheet_df.to_dict(orient="records")
            data.extend(records)

    # Handle case where 'Sr No' might be missing
    if data and 'Sr No' not in data[0]:
        for i, item in enumerate(data):
            item['Sr No'] = i + 1

    # Group data by Source PDF to match Excel format
    if data:
        # Ensure data is properly formatted for display
        for item in data:
            # Ensure consistent capitalization for key certificate fields
            for field in ['PLATE_NO', 'HEAT_NO', 'TEST_CERT_NO']:
                if field in item and item[field]:
                    item[field] = str(item[field]).upper().strip()
    
    # Count unique entries by certificate combination (instead of counting every row)
    unique_combinations = set()
    for item in data:
        combo = (
            item.get('PLATE_NO', ''), 
            item.get('HEAT_NO', ''), 
            item.get('TEST_CERT_NO', '')
        )
        unique_combinations.add(combo)
    
    context = {
        "data": data[::-1],  # latest first
        "total_pdfs": len(set([d.get("Source PDF", "") for d in data if d.get("Source PDF")])),
        "total_extracted": len(unique_combinations),  # Count unique combinations only
        "messages": messages.get_messages(request),
        "now": timezone.now(),
    }
    return render(request, "extractor/dashboard.html", context)

def upload_pdf(request):
    vendors = Vendor.objects.annotate(pdf_count=Count('pdfs')).order_by('name')
    
    if request.method == "POST":
        try:
            logger.info("Processing PDF upload request")
            vendor_id = request.POST.get("vendor")
            uploaded_file = request.FILES.get("pdf")
            
            # Validation
            if not vendor_id:
                messages.error(request, "Please select a vendor")
                return redirect("dashboard")
            if not uploaded_file:
                messages.error(request, "Please select a PDF file")
                return redirect("dashboard")
            if not uploaded_file.name.lower().endswith('.pdf'):
                messages.error(request, "Only PDF files are allowed")
                return redirect("dashboard")
            
            # Get vendor
            try:
                vendor = Vendor.objects.get(id=vendor_id)
            except Vendor.DoesNotExist:
                messages.error(request, "Invalid vendor selected")
                return redirect("dashboard")
            
            # === Improved duplicate check: compare file size and name pattern ===
            file_size = uploaded_file.size
            file_base_name = os.path.splitext(os.path.basename(uploaded_file.name))[0].split('_')[0]  # Strip suffixes
            
            # Flag to track if this is a duplicate file
            is_duplicate = False
            
            # Look for existing PDFs with similar name and same size
            existing_pdfs = UploadedPDF.objects.filter(file_size=file_size)
            for pdf in existing_pdfs:
                existing_name = os.path.splitext(os.path.basename(pdf.file.name))[0].split('_')[0]
                if file_base_name == existing_name:
                    # Clear any previous messages that might be causing confusion
                    storage = messages.get_messages(request)
                    storage.used = True
                    
                    # Add a more prominent duplicate warning with clear details
                    duplicate_msg = f"⚠️ DUPLICATE FILE DETECTED: '{uploaded_file.name}' appears to be a duplicate of an existing file. Upload canceled."
                    messages.error(request, duplicate_msg)
                    is_duplicate = True
                    break
                    
            # If duplicate detected, redirect to dashboard instead of upload page
            if is_duplicate:
                return redirect("dashboard")
            
            # Save uploaded PDF with size information
            uploaded_pdf = UploadedPDF.objects.create(
                vendor=vendor,
                file=uploaded_file,
                file_size=file_size
            )
            
            # Create necessary directories
            media_root = settings.MEDIA_ROOT
            upload_dir = os.path.join(media_root, 'uploads')
            extracted_dir = os.path.join(media_root, 'extracted')
            os.makedirs(upload_dir, exist_ok=True)
            os.makedirs(extracted_dir, exist_ok=True)

            # Load vendor config
            try:
                vendor_name = vendor.name.split()[0].lower()
                vendor_config_path = os.path.join(settings.BASE_DIR, 'extractor', 'vendor_configs', f"{vendor_name}_steel.json")
                media_config_path = os.path.join(settings.MEDIA_ROOT, 'vendor_configs', f"{vendor_name}_steel.json")
                
                if os.path.exists(vendor_config_path):
                    config_path = vendor_config_path
                elif os.path.exists(media_config_path):
                    config_path = media_config_path
                else:
                    raise ValidationError(f"Vendor configuration file not found for {vendor.name}")
                
                with open(config_path, 'r') as f:
                    vendor_config = json.load(f)
                
                # Validate config structure
                required_fields = ['vendor_id', 'vendor_name', 'fields']
                missing_fields = [field for field in required_fields if field not in vendor_config]
                if missing_fields:
                    raise ValidationError(f"Invalid vendor config: missing {', '.join(missing_fields)}")
                
                required_field_patterns = ['PLATE_NO', 'HEAT_NO', 'TEST_CERT_NO']
                missing_patterns = [field for field in required_field_patterns if field not in vendor_config['fields']]
                if missing_patterns:
                    raise ValidationError(f"Invalid vendor config: missing patterns for {', '.join(missing_patterns)}")
                
            except (json.JSONDecodeError, ValidationError) as e:
                messages.error(request, f"PDF could not be processed: {str(e)}")
                return redirect("dashboard")
            except Exception as e:
                messages.error(request, "PDF could not be scanned or matched. OCR fallback under development.")
                logger.error(f"OCR/Extraction failed: {str(e)}", exc_info=True)
                return redirect("dashboard")
            
            # Queue the extraction task
            try:
                task = process_pdf_file.delay(uploaded_pdf.id, vendor_config)
                
                # Store task ID in session
                request.session['last_task_id'] = task.id
                
                messages.success(request, "PDF uploaded successfully. Processing started in background.")
                return redirect("dashboard")
            except Exception as e:
                messages.error(request, "PDF could not be processed. Task queue failed.")
                logger.error(f"Task queue failed: {str(e)}", exc_info=True)
                return redirect("dashboard")

        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {str(e)}")
            return redirect("dashboard")

    return render(request, "extractor/upload.html", {"vendors": vendors})



# === NEW: Task status endpoint for progress bar ===
def task_status(request, task_id: str):
    res = AsyncResult(task_id)
    payload = {"state": res.state}
    if res.state == "PROGRESS":
        meta = res.info or {}
        payload.update({
            "current": meta.get("current", 0),
            "total": meta.get("total", 1),
            "phase": meta.get("phase", "")
        })
    elif res.state == "SUCCESS":
        payload.update(res.result or {})
        # Clear task ID from session when completed
        if 'last_task_id' in request.session and request.session['last_task_id'] == task_id:
            del request.session['last_task_id']
            request.session.modified = True
    elif res.state == "FAILURE":
        payload.update({"status": "failed", "message": "Task failed"})
        # Clear task ID from session when failed
        if 'last_task_id' in request.session and request.session['last_task_id'] == task_id:
            del request.session['last_task_id']
            request.session.modified = True
    return JsonResponse(payload)

# === NEW: Clear task ID from session ===
def clear_task_id(request):
    if 'last_task_id' in request.session:
        del request.session['last_task_id']
        request.session.modified = True
    return JsonResponse({"success": True})


# === NEW: Download & Master backup ===
def _save_master_backup(df: pd.DataFrame):
    """
    Save a master Excel with date-wise sheets, e.g., 2025-08-30, 2025-08-31...
    Saved under MEDIA_ROOT/backups/master.xlsx
    """
    backups_dir = os.path.join(settings.MEDIA_ROOT, "backups")
    os.makedirs(backups_dir, exist_ok=True)
    filename = os.path.join(backups_dir, "master.xlsx")
    sheet_name = timezone.localdate().isoformat()
    
    # Ensure we have all the required columns in the correct order
    columns = [
        'Sr No', 'Vendor', 'PLATE_NO', 'HEAT_NO', 'TEST_CERT_NO', 
        'Filename', 'Page', 'Source PDF', 'Created', 'Hash', 'Remarks'
    ]
    
    # Add Sr No if missing
    if 'Sr No' not in df.columns:
        df.insert(0, 'Sr No', range(1, len(df) + 1))
    
    # Reorder columns if needed
    available_columns = [col for col in columns if col in df.columns]
    df = df[available_columns + [col for col in df.columns if col not in columns]]
    
    # Replace NaN values with empty strings
    df = df.fillna("")

    if os.path.exists(filename):
        try:
            book = load_workbook(filename)
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.book = book
                # If sheet exists, append under it; if not, create new.
                startrow = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=(startrow == 0), startrow=startrow)
        except Exception as e:
            # Log error and fallback to writing fresh file
            logger.error(f"Error appending to Excel: {str(e)}", exc_info=True)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

from django.http import FileResponse

def download_excel(request):
    """
    Provides a downloadable Excel file with all extracted data.
    This file matches the format of the locally stored master.xlsx file.
    """
    backups_dir = os.path.join(settings.MEDIA_ROOT, "backups")
    master_path = os.path.join(backups_dir, "master.xlsx")

    if not os.path.exists(master_path):
        messages.error(request, "No extracted data available for download")
        return redirect("dashboard")

    try:
        # Return the file as a download attachment
        response = FileResponse(
            open(master_path, "rb"),
            as_attachment=True,
            filename="extracted_data.xlsx"
        )
        return response
    except Exception as e:
        logger.error(f"Error downloading Excel: {str(e)}", exc_info=True)
        messages.error(request, "Could not generate Excel file")
        return redirect("dashboard")
